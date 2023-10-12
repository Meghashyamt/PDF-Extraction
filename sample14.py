# # from PyPDF2 import PdfReader
# # import openpyxl

# # def get_sheet_names_from_excel(excel_file_path):
# #     """Reads and returns sheet names from the given Excel file."""
# #     wb = openpyxl.load_workbook(excel_file_path)
# #     return wb.sheetnames

# # def search_pdf_and_extract_nearby_text(pdf_file_path, keyword_mapping, num_lines=40, ignore_pages=None):
# #     """Searches all pages of the given PDF document for the given keyword and extracts nearby text.

# #     Args:
# #         pdf_file_path: The path to the PDF file.
# #         keyword_mapping: A dictionary mapping sheet names to keywords.
# #         num_lines: Number of lines to extract around the keyword (default is 40).
# #         ignore_pages: List of page numbers to ignore during extraction.

# #     Returns:
# #         A string containing the nearby text of the keyword found in any page.
# #     """
# #     # Open the PDF file.
# #     pdf_reader = PdfReader(pdf_file_path)

# #     nearby_text = ""
# #     # Iterate through all the pages of the PDF.
# #     for page_num, page in enumerate(pdf_reader.pages, start=1):
# #         if ignore_pages and page_num in ignore_pages:
# #             continue

# #         # Extract the text from the current page.
# #         text = page.extract_text()
# #         # Split the text into lines.
# #         lines = text.split("\n")
# #         # Iterate through keyword_mapping dictionary and search for keywords in the PDF.
# #         for sheet_name, keyword in keyword_mapping.items():
# #             keyword_line_indices = [i for i in range(len(lines)) if keyword in lines[i]]
# #             # Extract nearby text for each occurrence of the keyword on the page.
# #             for keyword_line_index in keyword_line_indices:
# #                 start_index = max(0, keyword_line_index - num_lines // 2)
# #                 end_index = min(len(lines), keyword_line_index + num_lines // 2 + 1)
# #                 nearby_text += f"Sheet Name: {sheet_name}\n"  # Add sheet name to the output.
# #                 nearby_text += " ".join(lines[start_index:end_index]) + "\n"

# #     return nearby_text

# # # Example usage:
# # excel_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\G-Excel.xlsx"
# # pdf_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\RFP-214.pdf"
# # ignore_pages = [9, 28, 29, 159]

# # # Define the mapping of sheet names to keywords.
# # keyword_mapping = {
# #     "Scope of Work": "Overview of the AMISP Scope of Work",
# #     #"Another Sheet": "Another Keyword",
# #     # Add more mappings as needed.
# # }

# # # Get sheet names from the Excel file.
# # sheet_names = get_sheet_names_from_excel(excel_file_path)

# # # Filter keyword mapping based on available sheet names.
# # filtered_keyword_mapping = {sheet_name: keyword_mapping[sheet_name] for sheet_name in sheet_names if sheet_name in keyword_mapping}

# # nearby_text = search_pdf_and_extract_nearby_text(pdf_file_path, filtered_keyword_mapping, num_lines=40, ignore_pages=ignore_pages)

# # print(nearby_text)

# from PyPDF2 import PdfReader
# import openpyxl

# def get_sheet_names_from_excel(excel_file_path):
#     """Reads and returns sheet names from the given Excel file."""
#     wb = openpyxl.load_workbook(excel_file_path)
#     return wb

# def search_pdf_and_extract_nearby_text(pdf_file_path, keyword_mapping, num_lines=40, ignore_pages=None):
#     """Searches all pages of the given PDF document for the given keyword and extracts nearby text.

#     Args:
#         pdf_file_path: The path to the PDF file.
#         keyword_mapping: A dictionary mapping sheet names to keywords.
#         num_lines: Number of lines to extract around the keyword (default is 40).
#         ignore_pages: List of page numbers to ignore during extraction.

#     Returns:
#         A dictionary containing the nearby text of the keyword found in any page.
#     """
#     # Open the PDF file.
#     pdf_reader = PdfReader(pdf_file_path)

#     nearby_text_dict = {}  # Dictionary to store nearby text for each sheet name.
    
#     # Iterate through all the pages of the PDF.
#     for page_num, page in enumerate(pdf_reader.pages, start=1):
#         if ignore_pages and page_num in ignore_pages:
#             continue

#         # Extract the text from the current page.
#         text = page.extract_text()
#         # Split the text into lines.
#         lines = text.split("\n")
#         # Iterate through keyword_mapping dictionary and search for keywords in the PDF.
#         for sheet_name, keyword in keyword_mapping.items():
#             keyword_line_indices = [i for i in range(len(lines)) if keyword in lines[i]]
#             nearby_text = ""
#             # Extract nearby text for each occurrence of the keyword on the page.
#             for keyword_line_index in keyword_line_indices:
#                 start_index = max(0, keyword_line_index - num_lines // 2)
#                 end_index = min(len(lines), keyword_line_index + num_lines // 2 + 1)
#                 nearby_text += " ".join(lines[start_index:end_index]) + "\n"

#             # Add nearby text to the dictionary with sheet name as key.
#             if sheet_name not in nearby_text_dict:
#                 nearby_text_dict[sheet_name] = ""
#             nearby_text_dict[sheet_name] += nearby_text

#     return nearby_text_dict


# # Example usage:
# excel_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\G-Excel.xlsx"
# pdf_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\RFP-214.pdf"
# ignore_pages = [9, 28, 29, 159]

# # Define the mapping of sheet names to keywords.
# keyword_mapping = {
#     "Scope of Work": "Overview of the AMISP Scope of Work",
#     "PQR":"Eligibility and Qualification Requirements"
#     # Add more mappings as needed.
# }

# # Get sheet names and workbook from the Excel file.
# wb = get_sheet_names_from_excel(excel_file_path)
# sheet_names = wb.sheetnames

# # Filter keyword mapping based on available sheet names.
# filtered_keyword_mapping = {sheet_name: keyword_mapping[sheet_name] for sheet_name in sheet_names if sheet_name in keyword_mapping}

# # Search for nearby text in the PDF.
# nearby_text_dict = search_pdf_and_extract_nearby_text(pdf_file_path, filtered_keyword_mapping, num_lines=40, ignore_pages=ignore_pages)

# # Write nearby text to corresponding sheets in the Excel file.
# for sheet_name, nearby_text in nearby_text_dict.items():
#     sheet = wb[sheet_name]
#     sheet['A1'] = nearby_text  # Write nearby text to cell A1, you can modify the destination cell as needed.

# # Save the modified workbook.
# wb.save(excel_file_path)

# print("Nearby text has been stored in the Excel file.")




# from PyPDF2 import PdfReader
# import openpyxl

# def get_sheet_names_from_excel(excel_file_path):
#     """Reads and returns sheet names from the given Excel file."""
#     wb = openpyxl.load_workbook(excel_file_path)
#     return wb

# def search_pdf_and_extract_nearby_text(pdf_file_path, keyword_mapping, num_lines=60, ignore_pages=None):
#     """Searches all pages of the given PDF document for the given keyword and extracts nearby text.

#     Args:
#         pdf_file_path: The path to the PDF file.
#         keyword_mapping: A dictionary mapping sheet names to keywords.
#         num_lines: Number of lines to extract around the keyword (default is 40).
#         ignore_pages: List of page numbers to ignore during extraction.

#     Returns:
#         A dictionary containing the nearby text of the keyword found in any page.
#     """
#     # Open the PDF file.
#     pdf_reader = PdfReader(pdf_file_path)

#     nearby_text_dict = {}  # Dictionary to store nearby text for each sheet name.
    
#     # Iterate through all the pages of the PDF.
#     for page_num, page in enumerate(pdf_reader.pages, start=1):
#         if ignore_pages and page_num in ignore_pages:
#             continue

#         # Extract the text from the current page.
#         text = page.extract_text()
#         # Split the text into paragraphs.
#         paragraphs = text.split("\n\n")  # Adjust the delimiter based on your PDF content.
#         # Iterate through keyword_mapping dictionary and search for keywords in the PDF.
#         for sheet_name, keyword in keyword_mapping.items():
#             keyword_line_indices = [i for i, para in enumerate(paragraphs) if keyword in para]
#             nearby_text = ""
#             # Extract nearby text for each occurrence of the keyword in paragraphs.
#             for keyword_line_index in keyword_line_indices:
#                 start_index = max(0, keyword_line_index - num_lines // 2)
#                 end_index = min(len(paragraphs), keyword_line_index + num_lines // 2 + 1)
#                 nearby_text += "\n\n".join(paragraphs[start_index:end_index]) + "\n\n"

#             # Add nearby text to the dictionary with sheet name as key.
#             if sheet_name not in nearby_text_dict:
#                 nearby_text_dict[sheet_name] = []
#             nearby_text_dict[sheet_name].append(nearby_text)

#     return nearby_text_dict

# # Example usage:
# excel_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\G-Excel.xlsx"
# pdf_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\RFP-214.pdf"
# ignore_pages = [9, 28, 29, 159]

# # Define the mapping of sheet names to keywords.
# keyword_mapping = {
#     "Scope of Work": "Overview of the AMISP Scope of Work",
#     # Add more mappings as needed.
# }

# # Get sheet names and workbook from the Excel file.
# wb = get_sheet_names_from_excel(excel_file_path)
# sheet_names = wb.sheetnames

# # Filter keyword mapping based on available sheet names.
# filtered_keyword_mapping = {sheet_name: keyword_mapping[sheet_name] for sheet_name in sheet_names if sheet_name in keyword_mapping}

# # Search for nearby text in the PDF.
# nearby_text_dict = search_pdf_and_extract_nearby_text(pdf_file_path, filtered_keyword_mapping, num_lines=40, ignore_pages=ignore_pages)

# # Write nearby text to corresponding sheets in the Excel file.
# for sheet_name, paragraphs in nearby_text_dict.items():
#     sheet = wb[sheet_name]
#     for i, paragraph in enumerate(paragraphs):
#         row = i + 1  # Start from row 1
#         sheet.cell(row=row, column=1, value=paragraph)  # Write paragraph to the corresponding row.

# # Save the modified workbook.
# wb.save(excel_file_path)

# print("Nearby text has been stored in the Excel file.")

# from PyPDF2 import PdfReader
# import openpyxl

# def get_sheet_names_from_excel(excel_file_path):
#     """Reads and returns sheet names from the given Excel file."""
#     wb = openpyxl.load_workbook(excel_file_path)
#     return wb

# def search_pdf_and_extract_nearby_text(pdf_file_path, keyword_mapping, num_lines=60, ignore_pages=None):
#     """Searches all pages of the given PDF document for the given keyword and extracts nearby text.

#     Args:
#         pdf_file_path: The path to the PDF file.
#         keyword_mapping: A dictionary mapping sheet names to keywords.
#         num_lines: Number of lines to extract around the keyword (default is 40).
#         ignore_pages: List of page numbers to ignore during extraction.

#     Returns:
#         A dictionary containing the nearby text of the keyword found in any page.
#     """
#     # Open the PDF file.
#     pdf_reader = PdfReader(pdf_file_path)

#     nearby_text_dict = {}  # Dictionary to store nearby text for each sheet name.
    
#     # Iterate through all the pages of the PDF.
#     for page_num, page in enumerate(pdf_reader.pages, start=1):
#         if ignore_pages and page_num in ignore_pages:
#             continue

#         # Extract the text from the current page.
#         text = page.extract_text()
#         # Split the text into paragraphs.
#         paragraphs = text.split("\n\n")  # Adjust the delimiter based on your PDF content.
#         # Iterate through keyword_mapping dictionary and search for keywords in the PDF.
#         for sheet_name, keyword in keyword_mapping.items():
#             keyword_line_indices = [i for i, para in enumerate(paragraphs) if keyword in para]
#             nearby_text = ""
#             # Extract nearby text for each occurrence of the keyword in paragraphs.
#             for keyword_line_index in keyword_line_indices:
#                 start_index = max(0, keyword_line_index - num_lines // 2)
#                 end_index = min(len(paragraphs), keyword_line_index + num_lines // 2 + 1)
#                 nearby_text += "\n\n".join(paragraphs[start_index:end_index]) + "\n\n"

#             # Add nearby text to the dictionary with sheet name as key.
#             if sheet_name not in nearby_text_dict:
#                 nearby_text_dict[sheet_name] = []
#             nearby_text_dict[sheet_name].append(nearby_text)

#     return nearby_text_dict

# # Example usage:
# excel_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\G-Excel.xlsx"
# pdf_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\RFP-214.pdf"
# ignore_pages = [9, 28, 29, 159]

# # Define the mapping of sheet names to keywords.
# keyword_mapping = {
#     "Scope of Work": "Overview of the AMISP Scope of Work",
#     # Add more mappings as needed.
# }

# # Get sheet names and workbook from the Excel file.
# wb = get_sheet_names_from_excel(excel_file_path)
# sheet_names = wb.sheetnames

# # Filter keyword mapping based on available sheet names.
# filtered_keyword_mapping = {sheet_name: keyword_mapping[sheet_name] for sheet_name in sheet_names if sheet_name in keyword_mapping}

# # Search for nearby text in the PDF.
# nearby_text_dict = search_pdf_and_extract_nearby_text(pdf_file_path, filtered_keyword_mapping, num_lines=40, ignore_pages=ignore_pages)

# # Write nearby text to corresponding sheets in the Excel file.
# for sheet_name, paragraphs in nearby_text_dict.items():
#     sheet = wb[sheet_name]
#     for i, paragraph in enumerate(paragraphs):
#         row = i + 1  # Start from row 1
#         sheet.cell(row=row, column=1, value=paragraph)  # Write paragraph to the corresponding row.

# # Save the modified workbook.
# wb.save(excel_file_path)

# print("Nearby text has been stored in the Excel file.")


import torch
from transformers import BertForQuestionAnswering, BertTokenizer

from PyPDF2 import PdfReader
import openpyxl

def get_sheet_names_from_excel(excel_file_path):
    """Reads and returns sheet names from the given Excel file."""
    wb = openpyxl.load_workbook(excel_file_path)
    return wb

def search_pdf_and_extract_nearby_text(pdf_file_path, keyword_mapping, num_lines=40, ignore_pages=None):
    # Load pre-trained BERT model and tokenizer
    model_name = 'bert-large-uncased-whole-word-masking-finetuned-squad' # You can choose other BERT variants if needed
    tokenizer = BertTokenizer.from_pretrained(model_name)
    model = BertForQuestionAnswering.from_pretrained(model_name)

    nearby_text_dict = {}  # Dictionary to store nearby text for each sheet name.

    # Iterate through all the pages of the PDF.
    for page_num, page in enumerate(PdfReader(pdf_file_path).pages, start=1):
        if ignore_pages and page_num in ignore_pages:
            continue

        # Extract the text from the current page.
        text = page.extract_text()
        # Split the text into lines.
        lines = text.split("\n")
        
        for sheet_name, keyword in keyword_mapping.items():
            nearby_text = ""
            # Iterate through lines and extract text using BERT-based question answering model.
            for line in lines:
                if keyword in line:
                    # Use BERT to extract relevant text around the keyword.
                    inputs = tokenizer.encode_plus(question=keyword, context=line, return_tensors="pt")
                    start_positions, end_positions = model(**inputs).values()
                    start_index = torch.argmax(start_positions, dim=1).item()
                    end_index = torch.argmax(end_positions, dim=1).item()
                    nearby_text += tokenizer.decode(inputs["input_ids"][0][start_index:end_index+1]) + "\n"

            # Add nearby text to the dictionary with sheet name as key.
            if sheet_name not in nearby_text_dict:
                nearby_text_dict[sheet_name] = ""
            nearby_text_dict[sheet_name] += nearby_text

    return nearby_text_dict

#Rest of your code remains unchanged.

# Example usage:
excel_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\G-Excel.xlsx"
pdf_file_path = r"C:\Users\M_Thiruveedula\Downloads\Projects_Code_Backup\Genus\RFP-214.pdf"
ignore_pages = [9, 28, 29, 159]

# Define the mapping of sheet names to keywords.
keyword_mapping = {
    "Scope of Work": "Overview of the AMISP Scope of Work",
    # Add more mappings as needed.
}

# Get sheet names and workbook from the Excel file.
wb = get_sheet_names_from_excel(excel_file_path)
sheet_names = wb.sheetnames

# Filter keyword mapping based on available sheet names.
filtered_keyword_mapping = {sheet_name: keyword_mapping[sheet_name] for sheet_name in sheet_names if sheet_name in keyword_mapping}

# Search for nearby text in the PDF.
nearby_text_dict = search_pdf_and_extract_nearby_text(pdf_file_path, filtered_keyword_mapping, num_lines=40, ignore_pages=ignore_pages)

# Write nearby text to corresponding sheets in the Excel file.
for sheet_name, paragraphs in nearby_text_dict.items():
    sheet = wb[sheet_name]
    for i, paragraph in enumerate(paragraphs):
        row = i + 1  # Start from row 1
        sheet.cell(row=row, column=1, value=paragraph)  # Write paragraph to the corresponding row.

# Save the modified workbook.
wb.save(excel_file_path)

print("Nearby text has been stored in the Excel file.")
